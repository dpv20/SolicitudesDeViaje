import streamlit as st
import pandas as pd
from openpyxl import load_workbook
import shutil
from typing import List
import datetime
import os
import win32com.client as win32
import pythoncom
from PyPDF2 import PdfWriter, PdfReader, PdfMerger
from functions_webpage import *
import datetime
import secrets
import string
import openpyxl
import time
from send_mail import *


countries_list = ["Afghanistan", "Alabama", "Akrotiri", "Albania", "Algeria", "Andorra", "Angola", "Anguilla", "Antigua and Barbuda", "Argentina", "Armenia", "Aruba", "Australia", "Austria", "Azerbaijan", "Bahamas", 
                 "Bahrain", "Barbados", "Belarus", "Belgium", "Belize", "Benin", "Bolivia", "Bosnia and Herzegovina", "Botswana", "Brazil", "British Virgin Islands", "Brunei", "Bulgaria", "Burkina Faso", "Burundi", 
                 "California", "Cambodia", "Cameroon", "Canada", "Cayman Islands", "Central African Republic", "Chad", "Chile", "China", "Colombia", "Comoros", "Congo", "Costa Rica", "Croatia", "Cuba", "Cyprus", 
                 "Czechia", "Democratic Republic of Congo", "Denmark", "Dhekelia", "Djibouti", "Djibouti", "Dominica", "Dominican Republic", "East Timor", "Ecuador", "Egypt", "El Salvador", "Equatorial Guinea", 
                 "Eritrea", "Eritrea", "Estonia", "Ethiopia", "Faroe Islands", "Fiji", "Finland", "Florida", "France", "French Guiana", "Gabon", "Galápagos Islands", "Gambia", "Georgia", "Germany", "Ghana", 
                 "Gibraltar", "Greece", "Grenada", "Guadeloupe", "Guatemala", "Guinea", "Guinea-Bissau", "Guyana", "Haiti", "Honduras", "Hong Kong", "Hungary", "Iceland", "Illinois", "India", "Indonesia", 
                 "Iran", "Iraq", "Ireland", "Israel", "Italy", "Ivory Coast", "Jamaica", "Japan", "Jordan", "Kazakhstan", "Kenya", "Kuwait", "Kyrgyzstan", "Laos", "Latvia", "Lebanon", "Lesotho", "Liberia", 
                 "Libya", "Liechtenstein", "Lithuania", "Louisiana", "Luxembourg", "Macau", "Macedonia", "Madagascar", "Malawi", "Malaysia", "Maldives", "Mali", "Malta", "Martinique", "Mauritania", "Mauritania", 
                 "Mauritius", "Mayotte", "Mississippi", "Mexico", "Moldova", "Monaco", "Mongolia", "Montenegro", "Montserrat", "Morocco", "Mozambique", "Myanmar", "Namibia", "Nepal", "Netherlands", 
                 "Netherlands Antilles", "New Guinea", "New Zealand", "Nicaragua", "Niger", "Niger", "Nigeria", "Norway", "Oman", "Pakistan", "Panama", "Paraguay", "Peru", "Philippines", "Poland", "Portugal", 
                 "Puerto Rico", "Qatar", "Reunion", "Romania", "Russia", "Rwanda", "Saint Barthélemy", "Saint Helena", "Saint Kits and Nevis", "Saint Lucia", "Saint Martin", "Saint Vincent and the Grenadines", 
                 "San Marino", "São tomé and Principe", "Saudi Arabia", "Senegal", "Serbia", "Seychelles", "Sierra Leone", "Singapore", "Slovakia", "Slovenia", "Somalia", "South Africa", "South Korea", 
                 "South Sudan", "Spain", "Sudan", "Sudan", "Suriname", "Suriname", "Swaziland", "Sweden", "Switzerland", "Syria", "Taiwan", "Tajikistan", "Tanzania", "Texas", "Thailand", "Togo", "Trinidad and Tobago", 
                 "Tunisia", "Turkey", "Turkmenistan", "Turks and Caicos Islands", "Uganda", "Ukraine", "United Arab Emirates", "United Kingdom", "United States Virgin Islands", "Uruguay", "Uzbekistan", "Venezuela", 
                 "Vietnam", "Wake Island", "Western Sahara", "Yemen", "Zambia"]

def calculate_machine_rooms(area):
    machine_rooms = {"XS": 0, "S": 0, "M": 0, "L": 0}
    
    while area > 0:
        if area >= 3.5:
            machine_rooms["L"] += 1
            area -= 7
        elif 3.5 >= area >= 1.5:
            machine_rooms["M"] += 1
            area -= 3.5
        elif 1.5 >= area >= 1:
            machine_rooms["S"] += 1
            area -= 1.5
        else:
            machine_rooms["XS"] += 1
            area -= 1
    
    return machine_rooms

def replace_formula_with_result(filename, decimal_separator):
    wb = openpyxl.load_workbook(filename, data_only=True)
    if decimal_separator == '.':
        thousands_separator = ','
    elif decimal_separator == ',':
        thousands_separator = '.'
    else:
        raise ValueError("Invalid decimal_separator: " + decimal_separator)

    for sheetname in wb.sheetnames:
        sheet = wb[sheetname]
        for row in sheet.iter_rows():
            for cell in row:
                if cell.data_type == 'f':  # f for formula
                    cell.value = cell.value  # replace the formula with its result
                
                # If cell contains a number, convert it to a string with the desired format
                if isinstance(cell.value, (int, float)):
                    # Check if the fractional part is zero
                    if cell.value % 1 == 0:
                        formatted_value = '{:,.0f}'.format(cell.value)  # format as integer
                    else:
                        formatted_value = '{:,.2f}'.format(cell.value)  # format as float

                    formatted_value = formatted_value.replace(",", "X").replace(".", "Y")
                    cell.value = formatted_value.replace("X", thousands_separator).replace("Y", decimal_separator)
    new_filename = 'new_' + filename.rsplit('.', 1)[0] + '.xlsx'  # change extension to .xlsx

    wb.save(new_filename)
    



def save_range_to_pdf(sheet, range_start, range_end, pdf_writer, wb):
    ws = wb.Worksheets(sheet)
    ws.PageSetup.Zoom = False
    ws.PageSetup.FitToPagesTall = 1
    ws.PageSetup.FitToPagesWide = 1

    rng = ws.Range(f"{range_start}:{range_end}")
    print(f"Attempting to export range: {range_start}:{range_end}")  # Add this line for debugging
    pdf_path = os.path.abspath(f"{sheet}_{range_start}_{range_end}.pdf")
    rng.ExportAsFixedFormat(0, pdf_path)

    with open(pdf_path, 'rb') as input_file:
        pdf_reader = PdfReader(input_file)
        for page in pdf_reader.pages:
            pdf_writer.add_page(page)

    
    os.remove(pdf_path)




def surveys(username):
    
    st.session_state['username']
    pythoncom.CoInitialize()
    st.title('Lagoon Design Form (Entire Project)')

    countries = ['None'] + countries_list
    regions = ['None', 'USA', 'ROW']
    lagoon_types = ['None', 'RE', 'PAL', 'HYBRID']
    decimal_separators = ['None', ',', '.']
    machine_room_options = ['None', 'Underground', 'Surface']
    languages = ['None', 'English', 'Spanish']
    currencies = ['None', 'USD', 'EUR', 'UF']
    TAGS =[]
    with st.form(key='lagoon_form'):
        project_name = st.text_input('Project Name', max_chars=50)
        decimal_separator = st.selectbox('Decimal Separator', decimal_separators)
        country = st.selectbox('Country', countries)
        area = st.number_input('Area', min_value=0.0, format="%.2f")
        region = st.selectbox('Region', regions)
        lagoon_type = st.selectbox('Type of Lagoon', lagoon_types)
        language = st.selectbox('Language', languages)
        currency = st.selectbox('Currency', currencies)

        calculated_machine_rooms = calculate_machine_rooms(area)

        max_depth = st.number_input('Maximum Lagoon Depth', min_value=0.0, format="%.2f")
        wall = st.number_input('Wall', key=f"wall", min_value=0.0, format="%.2f")
        wall_1_name = st.text_input('Wall 1 Name Type (optional)', max_chars=50, key='wall_1_name')
        wall_2 = st.number_input('Wall 2', key="wall 2", min_value=0.0, format="%.2f")
        wall_2_name = st.text_input('Wall 2 Name Type (optional)', max_chars=50, key='wall_2_name')
        wall_3 = st.number_input('Wall 3', key="wall 3", min_value=0.0, format="%.2f")
        wall_3_name = st.text_input('Wall 3 Name Type (optional)', max_chars=50, key='wall_3_name')
        beach_entrance = st.number_input('Beach Entrance', min_value=0.0, format="%.2f")
        island_beach_entrance = st.number_input('Island beach entrance', min_value=0.0, format="%.2f")
        machine_room = st.selectbox('Machine Room', machine_room_options)
        total_wall = wall + wall_2 + wall_3
        total_perimeter = total_wall + beach_entrance + island_beach_entrance 
        #usando las 3 en vez de 2 ahora



        uploaded_file = st.file_uploader("Choose a .dwg file", type="dwg")

        submit_button = st.form_submit_button('Submit')

        if submit_button:
            if (len(project_name) == 0 or  decimal_separator == 'None' or country == 'None' or region == 'None' or lagoon_type == 'None' or language == 'None' or currency == 'None' or machine_room == 'None'):
        
                st.error('Please make sure all fields are filled out.')

            else:
                answers = [
                    project_name, decimal_separator, country, area, region, lagoon_type, language, currency,
                    calculated_machine_rooms['XS'], calculated_machine_rooms['S'], calculated_machine_rooms['M'],
                    calculated_machine_rooms['L'], max_depth, total_perimeter, total_wall, beach_entrance,
                    island_beach_entrance, machine_room
                ]

                question_labels = [
                    'Project Name', 'Decimal Separator', 'Country', 'Area', 'Region', 'Type of Lagoon', 'Language', 'Currency',
                    'Machine Room - XS', 'Machine Room - S', 'Machine Room - M', 'Machine Room - L',
                    'Maximum Lagoon Depth', 'Total Perimeter', 'Wall', 'Beach Entrance', 'Island beach entrance', 'Machine Room'
                ]

                # Add the current date to the answers
                current_date = datetime.date.today().strftime('%Y/%m/%d')
                answers.append(current_date)
                answers.append('pendiente')
                answers.append(username)

                # Add the "Name" column with a random string
                random_string = generate_random_string()
                answers.append(random_string)

                question_labels.append('Date')
                question_labels.append('Estado')
                question_labels.append('Username')
                question_labels.append('TAG')

                if uploaded_file is not None:
                    file_name = "dwg/" + random_string + ".dwg"
                    with open(file_name, 'wb') as out:
                        out.write(uploaded_file.getbuffer())




                labeled_answers = {question: answer for question, answer in zip(question_labels, answers)}

                # Create a DataFrame with the answers
                data = pd.DataFrame([answers], columns=question_labels)

                # Append the data to the 'proyectos.csv' file
                with open('proyectos.csv', 'a') as f:
                    #data.to_csv(f, header=False, index=False, line_terminator='\n')
                    data.to_csv(f, header=False, index=False)


                # Now handle the excel output
                excel = win32.gencache.EnsureDispatch('Excel.Application')
                excel.Visible = False
                excel.DisplayAlerts = False
                wb = excel.Workbooks.Open(os.path.abspath('TEX.xlsm'))

                if lagoon_type == 'HYBRID':
                    lagoon_type = 'PAL'
                answer_cells = {
                    "H3": country,
                    "H4": area,
                    "H5": region,
                    "H6": lagoon_type,
                    "H8": calculated_machine_rooms['XS'],
                    "H9": calculated_machine_rooms['S'],
                    "H10": calculated_machine_rooms['M'],
                    "H11": calculated_machine_rooms['L'],
                    "H12": max_depth,
                    "H13": total_perimeter,
                    "H14": wall,
                    "H15": beach_entrance,
                    "H16": island_beach_entrance,
                    "H17": machine_room
                }

                ws = wb.Worksheets('Datos')
                for cell, answer in answer_cells.items():
                    ws.Range(cell).Value = answer

                # Save and close the workbook
                wb.Save()
                wb.Close()
                replace_formula_with_result('TEX.xlsm', decimal_separator)
                wb = excel.Workbooks.Open(os.path.abspath('new_TEX.xlsx'))
                # Create a new PDF writer
                pdf_writer = PdfWriter()


                # Call the save_range_to_pdf function based on the selected parameters
                if region == 'USA':
                    save_range_to_pdf('USA', 'B2', 'D84', pdf_writer, wb)
                elif region == 'ROW':
                    if language == 'English':
                        if currency == 'USD':
                            save_range_to_pdf('Template ROW', 'B3', 'M16', pdf_writer, wb)
                        elif currency == 'EUR':
                            save_range_to_pdf('Template ROW', 'B19', 'M32', pdf_writer, wb)
                    elif language == 'Spanish':
                        if currency == 'USD':
                            save_range_to_pdf('Template ESP', 'B4', 'J19', pdf_writer, wb)
                        elif currency == 'EUR':
                            save_range_to_pdf('Template ESP', 'B23', 'J39', pdf_writer, wb)
                        elif currency == 'UF':
                            save_range_to_pdf('Template UF', 'B3', 'J18', pdf_writer, wb)



                answers2 = answers
                question_labels2 = question_labels
                
                if wall: 
                    if wall_1_name:
                        answers2.append(wall_1_name)
                        question_labels2.append('Wall 1 Name Type')
                    answers2.append(wall)
                    question_labels2.append('Wall 1 size')

                if wall_2:
                    if wall_2_name:
                        answers2.append(wall_2_name)
                        question_labels2.append('Wall 2 Name Type')
                    answers2.append(wall_2)
                    question_labels2.append('Wall 2 size')

                if wall_3:
                    if wall_3_name:
                        answers2.append(wall_3_name)
                        question_labels2.append('Wall 3 Name Type')
                    answers2.append(wall_3)
                    question_labels2.append('Wall 3 size')
                
                labeled_answers2 = {question2: answer2 for question2, answer2 in zip(question_labels2, answers2)}

                st.write('Submitted answers:')
                a = "&nbsp;&nbsp;&nbsp;"
                excluded_labels = ['Machine Room - XS', 'Machine Room - S', 'Machine Room - M', 'Machine Room - L', 'TAG']
                for k, v in labeled_answers.items():
                    st.write(f'{k}:{a}{v}')

                os.makedirs(f'TEXs/{random_string}', exist_ok=True)

                name_txt = f'TEXs/{random_string}/{random_string}.txt'  
                with open(name_txt, 'w') as txt_file:
                    for k, v in labeled_answers2.items():
                        if k not in excluded_labels:
                            txt_file.write(f'{k}:   {v}\n')


                # Call the save_range_to_pdf function
                #save_range_to_pdf('Datos', 'H3', 'H17', pdf_writer, wb)
                
                
                
                # Create the "TEXs" directory if it doesn't exist
                os.makedirs(f'TEXs/{random_string}', exist_ok=True)

                # Then save your PDF
                with open(f'TEXs/{random_string}/{random_string}.pdf', 'wb') as output_pdf:
                    pdf_writer.write(output_pdf)



                TAGS.append(random_string)
                send_mail(f'{project_name} pendiente', TAGS, attach_pdf=False)

                time.sleep(1)

                st.success('Form submitted successfully.')


