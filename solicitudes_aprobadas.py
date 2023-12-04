import os
import streamlit as st
import pandas as pd
from pathlib import Path
from datetime import datetime
from pdf2image import convert_from_path
import base64
from openpyxl import load_workbook
from PIL import Image, ImageDraw, ImageFont
import os
from PyPDF2 import PdfFileWriter, PdfFileReader, PdfWriter, PdfReader
from reportlab.pdfgen import canvas
from fpdf import FPDF
from openpyxl import load_workbook
import win32com.client as win32
import tempfile
import pythoncom
import csv
import matplotlib.pyplot as plt
import unicodedata

def save_range_to_pdf(sheet, range_start, range_end, pdf_writer, wb, base_file_name):
    ws = wb.Worksheets(sheet)
    ws.PageSetup.Zoom = False
    ws.PageSetup.FitToPagesTall = 1
    ws.PageSetup.FitToPagesWide = 1
    rng = ws.Range(f"{range_start}:{range_end}")
    pdf_path = os.path.abspath(f"{base_file_name}.pdf")
    rng.ExportAsFixedFormat(0, pdf_path)
    with open(pdf_path, 'rb') as input_file:
        pdf_reader = PdfReader(input_file)
        for page in pdf_reader.pages:
            pdf_writer.add_page(page)
    os.remove(pdf_path)


def create_pdf_with_image_and_excel(base_file_name):
    pythoncom.CoInitialize()

    excel_path = os.path.join('formularios_viaje', f"{base_file_name}.xlsx")
    pdf_path = os.path.join('formularios_viaje', f"{base_file_name}.pdf")

    # Create a new PDF writer
    pdf_writer = PdfWriter()

    # Handle the Excel output
    excel = win32.gencache.EnsureDispatch('Excel.Application')
    excel.Visible = False
    excel.DisplayAlerts = False
    wb = excel.Workbooks.Open(os.path.abspath(excel_path))

    # Call the save_range_to_pdf function for the desired range
    save_range_to_pdf('Solicitud de Viaje', 'A1', 'K49', pdf_writer, wb, base_file_name)

    # Close the workbook
    wb.Close()

    # Save the final PDF
    with open(pdf_path, 'wb') as output_pdf:
        pdf_writer.write(output_pdf)


def update_total_usd_viatico(file_name, total_usd_viatico):
    script_dir = os.path.dirname(__file__)
    excel_file_path = os.path.join(script_dir, 'formularios_viaje', f"{file_name}.xlsx")

    if not os.path.exists(excel_file_path):
        print(f"Excel file not found: {excel_file_path}")
        return

    # Update the Excel file
    workbook = load_workbook(excel_file_path)
    sheet = workbook.active
    sheet['C47'] = f"{total_usd_viatico} USD"
    workbook.save(excel_file_path)

    # After updating the Excel file, call the function to create the PDF
    create_pdf_with_image_and_excel(file_name)


def modify_solicitud(selected_solicitud, new_data):
    file = 'solicitudes_viaje.csv'
    solicitudes = pd.read_csv(file)
    for column in new_data:
        if column != 'solicitud':
            solicitudes.loc[solicitudes['solicitud'] == selected_solicitud, column] = new_data[column]
    solicitudes.to_csv(file, index=False)

def solicitudes_aprobadas():
    st.title("Solicitudes de Viaje Aprobadas")
    file = 'solicitudes_viaje.csv'
    # Section control
    #section = st.radio("", ["Mostrar Listado", "Modificar Solicitud", "Eliminar Solicitud"])
    section = st.radio("Section", ["Mostrar Listado", "Modificar Solicitud", "Eliminar Solicitud"], label_visibility='collapsed')

    solicitudes = pd.read_csv(file)
    solicitudes.sort_values(by='solicitud', ascending=False, inplace=True)


    if section == "Mostrar Listado":
        # Read CSV file
        df_original = pd.read_csv("solicitudes_viaje.csv")
        df_original['fecha_solicitud'] = pd.to_datetime(df_original['fecha_solicitud'])
        df_original['fecha_partida'] = pd.to_datetime(df_original['fecha_partida'])
        df_original['fecha_llegada'] = pd.to_datetime(df_original['fecha_llegada'])

        # Filter for approved requests
        df_solicitudes = df_original[df_original['estado'] == 'aprobado']

        # Display table headers for selected columns
        header_cols = st.columns([3, 3, 3, 3, 3, 3, 3])
        col_names = ["Solicitud", "Fecha", "Nombre", "Viático", "Partida", "Llegada", "Seleccionar"]
        for i, col_name in enumerate(col_names):
            header_cols[i].write(col_name)

        # Display table with approved requests
        for index, row in df_solicitudes.iterrows():
            cols = st.columns([3, 3, 3, 3, 3, 3, 3])
            cols[0].write(row['solicitud'])
            cols[1].write(row['fecha_solicitud'].strftime('%d/%m/%Y'))
            cols[2].write(row['Nombre Empleado'])
            cols[3].write(row['total_usd_viatico'])
            cols[4].write(row['fecha_partida'].strftime('%d/%m/%Y'))
            cols[5].write(row['fecha_llegada'].strftime('%d/%m/%Y'))
            
            # Select button for each request
            select_button = cols[-1].button("Select", key=str(row['solicitud']))
            if select_button:
                st.session_state['selected_solicitud'] = row['solicitud']
            st.write("---")
        
        st.write("---")

        # Handle selected request
        if 'selected_solicitud' in st.session_state:
            selected_solicitud = st.session_state['selected_solicitud']
            selected_data = df_original[df_original['solicitud'] == selected_solicitud].iloc[0]

            fecha_selected_solicitud = df_original.loc[df_original['solicitud'] == selected_solicitud, 'fecha_solicitud'].iloc[0]
            solicitud_filename = f"{fecha_selected_solicitud.strftime('%Y%m%d')}_{selected_solicitud}"
            if not pd.isna(selected_data['comentario']) and selected_data['comentario'] != '':
                st.subheader(f"**Comentario:** {selected_data['comentario']}")
            # Display the PDF associated with the request
            pdf_path = os.path.join('formularios_viaje', f"{solicitud_filename}.pdf")
            if os.path.exists(pdf_path):
                images = convert_from_path(pdf_path)
                for image in images:
                    st.image(image, width=1000)
            else:
                st.error("Error mostrando el PDF.")

            # Button to download the Excel file
            excel_path = os.path.join('formularios_viaje', f"{solicitud_filename}.xlsx")
            if os.path.exists(excel_path):
                with open(excel_path, 'rb') as file:
                    excel_data = base64.b64encode(file.read()).decode()
                    href = f'<a href="data:application/octet-stream;base64,{excel_data}" download="{solicitud_filename}.xlsx">Descargar Excel</a>'
                    st.markdown(href, unsafe_allow_html=True)
            else:
                st.error("Error al descargar el archivo Excel.")

            # Function to toggle image display
            def toggle_image_display():
                key = f'show_image_{selected_solicitud}'
                if key in st.session_state:
                    st.session_state[key] = not st.session_state[key]
                else:
                    st.session_state[key] = True

            # Button to display image
            img_path_png = os.path.join('formularios_viaje', f"{solicitud_filename}.png")
            img_path_jpg = os.path.join('formularios_viaje', f"{solicitud_filename}.jpg")
            
            if os.path.exists(img_path_png) or os.path.exists(img_path_jpg):
                img_path = img_path_png if os.path.exists(img_path_png) else img_path_jpg
                toggle_button = st.button("Mostrar/Ocultar Imagen", key=f'toggle_image_{selected_solicitud}', on_click=toggle_image_display)
                
                if st.session_state.get(f'show_image_{selected_solicitud}', False):
                    st.image(img_path, width=500)

    elif section == "Eliminar Solicitud":
            st.subheader("Eliminar Solicitud")
            solicitudes_pendiente = solicitudes[solicitudes['estado'] == 'aprobado']
            solicitudes_list = solicitudes_pendiente['solicitud'].tolist()
            
            #selected_solicitud = st.selectbox("Seleccione una solicitud para eliminar:", [""] + solicitudes_list)
            selected_solicitud = st.selectbox("Seleccione una solicitud", [""] + solicitudes_list, label_visibility='collapsed')

            if st.button("Eliminar"):
                if selected_solicitud:
                    solicitudes = solicitudes[solicitudes['solicitud'] != selected_solicitud]
                    solicitudes.to_csv(file, index=False)
                    st.success(f"Se ha eliminado la solicitud {selected_solicitud}.")
                else:
                    st.warning("Por favor, selecciona una solicitud para eliminar.")
    elif section == "Modificar Solicitud":
        st.subheader("Modificar Solicitud")
        solicitudes_pendiente = solicitudes[solicitudes['estado'] == 'aprobado']
        solicitudes_list = solicitudes_pendiente['solicitud'].tolist()
        selected_solicitud = st.selectbox("Seleccione una solicitud para modificar:", [""] + solicitudes_list)

        if selected_solicitud:
            df_original = pd.read_csv("solicitudes_viaje.csv")
            df_original['fecha_solicitud'] = pd.to_datetime(df_original['fecha_solicitud'])
            filtered_df = df_original[df_original['solicitud'] == selected_solicitud]

            if not filtered_df.empty:
                fecha_selected_solicitud = filtered_df['fecha_solicitud'].iloc[0]
                solicitud_filename = f"{fecha_selected_solicitud.strftime('%Y%m%d')}_{selected_solicitud}"
                
                selected_data = solicitudes[solicitudes['solicitud'] == selected_solicitud].iloc[0]

                st.write("Solicitud:", selected_data['solicitud'])
                total_usd_viatico = st.number_input("Total USD Viático:", value=selected_data['total_usd_viatico'])

                if st.button("Guardar Cambios"):
                    # Update the total_usd_viatico in the Excel file
                    update_total_usd_viatico(solicitud_filename, total_usd_viatico)

                    # Update the value in the DataFrame
                    solicitudes.loc[solicitudes['solicitud'] == selected_solicitud, 'total_usd_viatico'] = total_usd_viatico
                    solicitudes.to_csv(file, index=False)

                    st.success(f"Se ha modificado la solicitud {selected_solicitud}.")
            else:
                st.warning("No se encontró la solicitud seleccionada.")



    '''
    elif section == "Modificar Solicitud":
        st.subheader("Modificar Solicitud")
        
        # Filter the data to list only 'aprobado' estado
        solicitudes_pendiente = solicitudes[solicitudes['estado'] == 'aprobado']
        solicitudes_list = solicitudes_pendiente['solicitud'].tolist()
        
        selected_solicitud = st.selectbox("Seleccione una solicitud para modificar:", [""] + solicitudes_list)
        
        if selected_solicitud:
            selected_data = solicitudes[solicitudes['solicitud'] == selected_solicitud].iloc[0]
            
            # Solicitud number should not be modifiable
            st.write("Solicitud:", selected_data['solicitud'])
            
            # Fields that can be modified
            nombre_empleado = st.text_input("Nombre Empleado:", value=selected_data['Nombre Empleado'])
            departamento = st.text_input("Departamento:", value=selected_data['departamento'])
            total_usd_viatico = st.number_input("Total USD Viático:", value=selected_data['total_usd_viatico'])
            fecha_partida = st.date_input("Fecha de Partida:", value=pd.to_datetime(selected_data['fecha_partida']))
            fecha_llegada = st.date_input("Fecha de Llegada:", value=pd.to_datetime(selected_data['fecha_llegada']))
            comentario = st.text_area("Comentario:", value=selected_data['comentario'])
            
            if st.button("Guardar Cambios"):
                solicitudes.loc[solicitudes['solicitud'] == selected_solicitud, 'Nombre Empleado'] = nombre_empleado
                solicitudes.loc[solicitudes['solicitud'] == selected_solicitud, 'departamento'] = departamento
                solicitudes.loc[solicitudes['solicitud'] == selected_solicitud, 'total_usd_viatico'] = total_usd_viatico
                solicitudes.loc[solicitudes['solicitud'] == selected_solicitud, 'fecha_partida'] = fecha_partida
                solicitudes.loc[solicitudes['solicitud'] == selected_solicitud, 'fecha_llegada'] = fecha_llegada
                solicitudes.loc[solicitudes['solicitud'] == selected_solicitud, 'comentario'] = comentario
                solicitudes.to_csv(file, index=False)
                st.success(f"Se ha modificado la solicitud {selected_solicitud}.")
    '''
