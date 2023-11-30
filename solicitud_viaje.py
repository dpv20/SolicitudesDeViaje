import streamlit as st
import pandas as pd
import unicodedata
from openpyxl import load_workbook
from datetime import datetime
import os
import matplotlib.pyplot as plt

from openpyxl import load_workbook
from PIL import Image, ImageDraw, ImageFont
from fpdf import FPDF
import os
import tempfile
import pythoncom
import win32com.client as win32
from PyPDF2 import PdfFileWriter, PdfFileReader, PdfWriter, PdfReader
from reportlab.pdfgen import canvas
import csv


def save_range_to_pdf(sheet, range_start, range_end, pdf_writer, wb):
    ws = wb.Worksheets(sheet)
    ws.PageSetup.Zoom = False
    ws.PageSetup.FitToPagesTall = 1
    ws.PageSetup.FitToPagesWide = 1
    rng = ws.Range(f"{range_start}:{range_end}")
    pdf_path = os.path.abspath(f"{sheet}_{range_start}_{range_end}.pdf")
    rng.ExportAsFixedFormat(0, pdf_path)
    with open(pdf_path, 'rb') as input_file:
        pdf_reader = PdfReader(input_file)
        for page in pdf_reader.pages:
            pdf_writer.add_page(page)
    os.remove(pdf_path)


''' #con imagen el PDF
def create_pdf_with_image_and_excel(numeronombre):
    pythoncom.CoInitialize()

    date_str = datetime.now().strftime('%Y%m%d')
    excel_path = os.path.join('formularios_viaje', f"{date_str}_{numeronombre}.xlsx")
    image_path = os.path.join('formularios_viaje', f"{date_str}_{numeronombre}.png")
    pdf_path = os.path.join('formularios_viaje', f"{date_str}_{numeronombre}.pdf")

    # Create a new PDF writer
    pdf_writer = PdfWriter()

    # Handle the Excel output
    excel = win32.gencache.EnsureDispatch('Excel.Application')
    excel.Visible = False
    excel.DisplayAlerts = False
    wb = excel.Workbooks.Open(os.path.abspath(excel_path))

    # Call the save_range_to_pdf function for the desired range
    save_range_to_pdf('Solicitud de Viaje', 'A1', 'K49', pdf_writer, wb)

    # Close the workbook
    wb.Close()

    # Add the image if it exists
    if os.path.exists(image_path):
        original_image = Image.open(image_path)
        temp_image_path = tempfile.mktemp(suffix=".pdf")
        
        # Create a new PDF with the dimensions of the image
        c = canvas.Canvas(temp_image_path, pagesize=(original_image.width, original_image.height))
        c.drawImage(image_path, 0, 0, original_image.width, original_image.height)
        c.save()
        original_image.close()
        
        with open(temp_image_path, 'rb') as input_file:
            pdf_reader = PdfReader(input_file)
            for page in pdf_reader.pages:
                pdf_writer.add_page(page)
        os.remove(temp_image_path)

    # Then save your PDF
    with open(pdf_path, 'wb') as output_pdf:
        pdf_writer.write(output_pdf)

    # Remove original Excel and image files
    #os.remove(excel_path)
    if os.path.exists(image_path):
        #os.remove(image_path)
        pass
'''


def create_pdf_with_image_and_excel(numeronombre): #sin imagen el pdf
    pythoncom.CoInitialize()

    date_str = datetime.now().strftime('%Y%m%d')
    excel_path = os.path.join('formularios_viaje', f"{date_str}_{numeronombre}.xlsx")
    pdf_path = os.path.join('formularios_viaje', f"{date_str}_{numeronombre}.pdf")

    # Create a new PDF writer
    pdf_writer = PdfWriter()

    # Handle the Excel output
    excel = win32.gencache.EnsureDispatch('Excel.Application')
    excel.Visible = False
    excel.DisplayAlerts = False
    wb = excel.Workbooks.Open(os.path.abspath(excel_path))

    # Call the save_range_to_pdf function for the desired range
    save_range_to_pdf('Solicitud de Viaje', 'A1', 'K49', pdf_writer, wb)

    # Close the workbook
    wb.Close()

    # Save the final PDF
    with open(pdf_path, 'wb') as output_pdf:
        pdf_writer.write(output_pdf)



def increment_number_in_file(file_path):
    # Leer el contenido del archivo
    with open(file_path, 'r') as file:
        number_str = file.read().strip()

    # Convertir el contenido a un número entero
    numero = int(number_str)

    # Incrementar el número
    numero += 1

    # Escribir el número incrementado en el archivo
    with open(file_path, 'w') as file:
        file.write(str(numero))

    print(f'Número leído: {numero - 1}')
    print(f'Número incrementado y guardado en el archivo: {numero}')

    return numero



def remove_accents(input_str):
    nfkd_form = unicodedata.normalize('NFKD', input_str)
    return u"".join([c for c in nfkd_form if not unicodedata.combining(c)])

def update_excel_file(respuesta, numeronombre):
    # Load the existing Excel file
    workbook = load_workbook('formulario_viaje.xlsx')
    sheet = workbook.active

    # Update the cells based on the given data
    sheet['H3'] = datetime.now().strftime('%d/%m/%Y')
    sheet['C6'] = respuesta.get('nombre_empleado', '')
    sheet['C8'] = respuesta.get('compania', '')
    sheet['G8'] = respuesta.get('cargo', '')
    sheet['C10'] = respuesta.get('departamento', '')
    sheet['C12'] = respuesta.get('tipo_viaje', '')
    sheet['G12'] = respuesta.get('transporte', '')
    sheet['C14'] = respuesta.get('clase', '')
    sheet['G14'] = respuesta.get('financiamiento', '')
    sheet['C19'] = respuesta.get('projects_string', '')
    sheet['C24'] = respuesta.get('proposito', '')
    sheet['C29'] = respuesta.get('condiciones_cobro', '')

    # Update the trayectos data starting from row 38
    trayectos_data = respuesta.get('trayectos', [])
    total_dias_viaje = 0
    total_usd_viatico = 0

    for i, trayecto in enumerate(trayectos_data, start=38):
        sheet[f'A{i}'] = trayecto['destino']
        sheet[f'B{i}'] = trayecto['fecha_partida'].strftime('%d/%m/%Y')
        sheet[f'C{i}'] = trayecto['lugar_partida']
        sheet[f'D{i}'] = trayecto['fecha_llegada'].strftime('%d/%m/%Y')
        sheet[f'E{i}'] = trayecto['lugar_llegada']
        sheet[f'F{i}'] = trayecto['financiamiento_trayecto']
        sheet[f'H{i}'] = trayecto['visa']
        numero_dias = trayecto['dias_totales']
        sheet[f'G{i}'] = numero_dias
        sheet[f'I{i}'] = ', '.join(trayecto['proyecto_trayecto'])  # Assuming proyecto_trayecto is a list
        sheet[f'J{i}'] = trayecto['Extras']

        total_dias_viaje += numero_dias
        total_usd_viatico += trayecto['usd_viatico']

    # Adding the requested values
    sheet['C16'] = trayectos_data[0]['fecha_partida'].strftime('%d/%m/%Y') if trayectos_data else ''
    sheet['G16'] = trayectos_data[-1]['fecha_llegada'].strftime('%d/%m/%Y') if trayectos_data else ''
    sheet['C47'] = f"{total_usd_viatico} USD"
    sheet['H47'] = total_dias_viaje


    # Save the updated workbook in the 'formularios_viaje' directory with a timestamp
    if not os.path.exists('formularios_viaje'):
        os.makedirs('formularios_viaje')
    save_path = os.path.join('formularios_viaje', f"{datetime.now().strftime('%Y%m%d')}_{numeronombre}.xlsx")
    workbook.save(save_path)


def solicitud(usuario):
    st.title('Solicitud de Viaje')
    st.write(f"{usuario}")

    # Load employee and project data
    df_empleados = pd.read_csv('datos_empleados.csv')
    df_proyectos = pd.read_csv('listado_proyectos.csv')
    available_projects = df_proyectos['Project Name'].tolist()
    project_codes = df_proyectos['Project Code'].tolist()
    
    # Extract the employee details if the email exists in the DataFrame
    empleado = df_empleados[df_empleados['email'] == usuario].iloc[0] if usuario in df_empleados['email'].values else None

    nombre_empleado_default = empleado['nombre completo'] if empleado is not None else ''
    compania_default = empleado['compania'] if empleado is not None else ''
    cargo_default = empleado['cargo'] if empleado is not None else ''
    departamento_default = empleado['departamento'] if empleado is not None else ''

    nombre_empleado = st.text_input('Nombre Empleado', value=nombre_empleado_default)
    compania_options = ['Crystal Lagoons BV', 'Crystal Lagoons Chile SpA', 'Inversiones Lagunas SpA', 'Crystal Lagoons US Corp']
    compania = st.selectbox('Compañía empleado', compania_options, index=compania_options.index(compania_default) if compania_default in compania_options else 0)
    cargo = st.text_input('Cargo', value=cargo_default)
    departamento_options = ['Administración y Finanzas', 'Arquitectura', 'Comercial', 'Development', 'Ingeniería', 'IT', 'Investigación y Desarrollo', 'Legal', 'Marketing', 'Operaciones', 'Servicios Generales', 'Water', 'Commercial', 'Engineering', 'Administration', 'Technical', 'Architecture', 'F&A', 'Management', 'R&D', 'Comptroller', 'HR', 'Legal']
    departamento = st.selectbox('Departamento', departamento_options, index=departamento_options.index(departamento_default) if departamento_default in departamento_options else 0)

    tipo_viaje_options = ['','Viaje Anual', 'Viaje de Negocios', 'Visita a Terreno', 'Visita Operacional', 'Otro']
    tipo_viaje = st.selectbox('Tipo de Viaje', tipo_viaje_options)
    tipo_viaje_otro = ''
    if tipo_viaje == 'Otro':
        tipo_viaje_otro = st.text_input('Especificar tipo de viaje')
    
    transporte_options = ['','Auto', 'Aéreo', 'Tren', 'Otro']
    transporte = st.selectbox('Medio de Transporte', transporte_options)
    transporte_otro = ''
    if transporte == 'Otro':
        transporte_otro = st.text_input('Especificar medio de transporte')

    clase_options = ['','Económica', 'Ejecutiva', 'Premium economy']
    clase = st.selectbox('Clase en que viaja', clase_options)

    financiamiento_options = compania_options + ['Cobrar al cliente', 'Cliente envía tickets', 'Otro']
    financiamiento = st.selectbox('Financiamiento', financiamiento_options, index=compania_options.index(compania))
    financiamiento_otro = ''
    if financiamiento == 'Otro':
        financiamiento_otro = st.text_input('Especificar financiamiento')

    selected_projects = st.multiselect('Selecciona los proyectos:', available_projects)
    selected_project_codes = [project_codes[available_projects.index(proj)] for proj in selected_projects]
    st.write('Códigos de proyectos seleccionados:', ', '.join(selected_project_codes))

    selected_projects_with_codes = [f"{proj} ({project_codes[available_projects.index(proj)]})" for proj in selected_projects]
    projects_string = ', '.join(selected_projects_with_codes)

    proposito = st.text_area('Propósito del Viaje')
    condiciones_cobro = st.text_area('Condiciones de cobro al cliente:')


    st.header("Detalle del viaje:")

    num_trayectos = st.slider('Número de trayectos', 2, 8)

    trayectos_data = []
    for i in range(0, num_trayectos, 2):
        cols = st.columns(2)
        
        with cols[0]:
            if i < num_trayectos:
                st.subheader(f'Trayecto {i+1}')
                trayectos_data.append(get_trayecto_data(i+1, selected_project_codes))
        
        with cols[1]:
            if i+1 < num_trayectos:
                st.subheader(f'Trayecto {i+2}')
                trayectos_data.append(get_trayecto_data(i+2, selected_project_codes))

    uploaded_image = st.file_uploader('Sube una imagen relacionada con el viaje:', type=['jpg', 'jpeg', 'png'])
    
    file_path_txt = 'formularios_viaje/numero_sol_viaje.txt'
    csv_path = 'solicitudes_viaje.csv'
    
    print(f"Ruta del archivo: {file_path_txt}")

    if st.button('Enviar'):
        numeronombre = increment_number_in_file(file_path_txt)
        nombre_solicitud = numeronombre
        fecha_solicitud = f"{datetime.now().strftime('%Y-%m-%d')}"
        print(fecha_solicitud)
        if uploaded_image:
            image_save_path = os.path.join('formularios_viaje', f"{datetime.now().strftime('%Y%m%d')}_{numeronombre}.png")
            with open(image_save_path, 'wb') as image_file:
                image_file.write(uploaded_image.read())
            st.write(f'Imagen guardada en: {image_save_path}')


        respuesta = {
            'nombre_empleado': nombre_empleado,
            'compania': compania,
            'cargo': cargo,
            'departamento': departamento,
            'tipo_viaje': tipo_viaje if tipo_viaje != 'Otro' else tipo_viaje_otro,
            'transporte': transporte if transporte != 'Otro' else transporte_otro,
            'clase': clase,
            'financiamiento': financiamiento if financiamiento != 'Otro' else financiamiento_otro,
            'proposito': proposito,
            'condiciones_cobro': condiciones_cobro,
            'proyectos': selected_projects,
            'codigos_proyectos': selected_project_codes,
            'trayectos': trayectos_data,
            'projects_string': projects_string
        }
        update_excel_file(respuesta, numeronombre)
        create_pdf_with_image_and_excel(numeronombre)

        with open(csv_path, 'a', newline='', encoding='utf-8') as csv_file:
            writer = csv.writer(csv_file)
            departamento = respuesta['departamento']
            total_usd_viatico = sum(trayecto['usd_viatico'] for trayecto in respuesta['trayectos'])
            fecha_partida = respuesta['trayectos'][0]['fecha_partida'].strftime('%Y-%m-%d') if respuesta['trayectos'] else ''
            fecha_llegada = respuesta['trayectos'][-1]['fecha_llegada'].strftime('%Y-%m-%d') if respuesta['trayectos'] else ''
            writer.writerow([nombre_solicitud, fecha_solicitud,'pendiente', nombre_empleado, departamento, total_usd_viatico, fecha_partida, fecha_llegada])


        st.write(respuesta)
        

def get_trayecto_data(index, project_codes):
    # Leer el archivo CSV
    viaticos_data = pd.read_csv('datos_viaticos.csv')
    destinos = [""] + viaticos_data.iloc[:, 0].tolist()
    usd_viaticos = viaticos_data.iloc[:, 1].tolist()

    destino = st.selectbox(f"Destino Trayecto {index}", destinos)
    fecha_partida = st.date_input(f"Fecha (Partida) Trayecto {index}")
    lugar_partida = st.text_input(f"Lugar (Partida) Trayecto {index}")
    fecha_llegada = st.date_input(f"Fecha (Llegada) Trayecto {index}")
    lugar_llegada = st.text_input(f"Lugar (Llegada) Trayecto {index}")
    visa = st.selectbox(f"Visa Trayecto {index}", ["", "Si", "No"])
    proyecto_trayecto = st.multiselect(f'Proyecto(s) Trayecto {index}', project_codes)
    financiamiento_trayecto = st.selectbox(f'Financiamiento Trayecto {index}', ["", 'Cobro al cliente', 'No cobro al cliente'])
    Extras = st.text_input(f"Nombre de hoteles de preferencia Trayecto {index}")

    # Obtener el valor de usd_viatico correspondiente al destino seleccionado
    usd_viatico_unitario = usd_viaticos[destinos.index(destino) - 1] if destino else 0
    numero_dias = (fecha_llegada - fecha_partida).days
    usd_viatico_total = numero_dias * usd_viatico_unitario if usd_viatico_unitario else 0

    return {
        'destino': destino,
        'fecha_partida': fecha_partida,
        'lugar_partida': lugar_partida,
        'fecha_llegada': fecha_llegada,
        'lugar_llegada': lugar_llegada,
        'visa': visa,
        'proyecto_trayecto': proyecto_trayecto,
        'financiamiento_trayecto': financiamiento_trayecto,
        'usd_viatico': usd_viatico_total,
        'Extras': Extras,
        'dias_totales': numero_dias
    }